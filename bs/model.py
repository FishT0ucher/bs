import numpy as np
import openpyxl
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor
from sklearn.neural_network import MLPRegressor
from sklearn.preprocessing import MinMaxScaler
from sklearn.tree import DecisionTreeRegressor
from scipy.stats import t
import torch
import torch.nn as nn
from torch import optim
import torch.nn.functional as F
from sklearn.metrics import mean_squared_error, mean_absolute_error, mean_squared_log_error, r2_score
mse_tree = []
mse_rf = []
mse_gbdt = []

file_path1 = 'D:\\M4_1\\bs\\data1.xlsx'  # 请替换为您的文件路径
workbook1 = openpyxl.load_workbook(file_path1)
sheet1 = workbook1.active

file_path2 = 'D:\\M4_1\\bs\\data2.xlsx'  # 请替换为您的文件路径
workbook2 = openpyxl.load_workbook(file_path2)
sheet2 = workbook2.active

file_path3 = 'D:\\M4_1\\bs\\data3.xlsx'  # 请替换为您的文件路径
workbook3 = openpyxl.load_workbook(file_path3)
sheet3 = workbook3.active

delay1 = [cell.value for cell in sheet1['A']]
snr1 = [cell.value for cell in sheet1['B']]
plr1 = [cell.value for cell in sheet1['C']]
delay1 = np.array(delay1[1:]).reshape(-1, 1)
snr1 = np.array(snr1[1:]).reshape(-1, 1)
plr1 = np.array(plr1[1:]).reshape(-1, 1)

delay2 = [cell.value for cell in sheet2['A']]
snr2 = [cell.value for cell in sheet2['B']]
plr2 = [cell.value for cell in sheet2['C']]
delay2 = np.array(delay2[1:]).reshape(-1, 1)
snr2 = np.array(snr2[1:]).reshape(-1, 1)
plr2 = np.array(plr2[1:]).reshape(-1, 1)

delay3 = [cell.value for cell in sheet3['A']]
snr3 = [cell.value for cell in sheet3['B']]
plr3 = [cell.value for cell in sheet3['C']]
plr0 = [cell.value for cell in sheet3['D']]
delay3 = np.array(delay3[1:]).reshape(-1, 1)
snr3 = np.array(snr3[1:]).reshape(-1, 1)
plr3 = np.array(plr3[1:]).reshape(-1, 1)
plr0 = np.array(plr0[1:]).reshape(-1, 1)


# 仅选择大于零的元素
positive_delay1 = delay1[delay1 > 0].reshape(-1, 1)
# 使用 MinMaxScaler 进行归一化
scaler_delay1 = MinMaxScaler()
normalized_positive_delay1 = scaler_delay1.fit_transform(positive_delay1)
# 将归一化后的结果填充回原始向量中
delay1[delay1 > 0] = normalized_positive_delay1.flatten()

# 仅选择大于零的元素
positive_snr1 = snr1[snr1 > 0].reshape(-1, 1)
# 使用 MinMaxScaler 进行归一化
scaler_snr1 = MinMaxScaler()
normalized_positive_snr1 = scaler_snr1.fit_transform(positive_snr1)
# 将归一化后的结果填充回原始向量中
snr1[snr1 > 0] = normalized_positive_snr1.flatten()

# 仅选择大于零的元素
positive_plr1 = plr1[plr1 > 0].reshape(-1, 1)
# 使用 MinMaxScaler 进行归一化
scaler_plr1 = MinMaxScaler()
normalized_positive_plr1 = scaler_plr1.fit_transform(positive_plr1)
# 将归一化后的结果填充回原始向量中
plr1[plr1 > 0] = normalized_positive_plr1.flatten()


# 仅选择大于零的元素
positive_delay2 = delay2[delay2 > 0].reshape(-1, 1)
# 使用 MinMaxScaler 进行归一化
scaler_delay2 = MinMaxScaler()
normalized_positive_delay2 = scaler_delay2.fit_transform(positive_delay2)
# 将归一化后的结果填充回原始向量中
delay2[delay2 > 0] = normalized_positive_delay2.flatten()

# 仅选择大于零的元素
positive_snr2 = snr2[snr2 > 0].reshape(-1, 1)
# 使用 MinMaxScaler 进行归一化
scaler_snr2 = MinMaxScaler()
normalized_positive_snr2 = scaler_snr2.fit_transform(positive_snr2)
# 将归一化后的结果填充回原始向量中
snr2[snr2 > 0] = normalized_positive_snr2.flatten()

# 仅选择大于零的元素
positive_plr2 = plr2[plr2 > 0].reshape(-1, 1)
# 使用 MinMaxScaler 进行归一化
scaler_plr2 = MinMaxScaler()
normalized_positive_plr2 = scaler_plr2.fit_transform(positive_plr2)
# 将归一化后的结果填充回原始向量中
plr2[plr2 > 0] = normalized_positive_plr2.flatten()


# 仅选择大于零的元素
positive_delay3 = delay3[delay3 > 0].reshape(-1, 1)
# 使用 MinMaxScaler 进行归一化
scaler_delay3 = MinMaxScaler()
normalized_positive_delay3 = scaler_delay3.fit_transform(positive_delay3)
# 将归一化后的结果填充回原始向量中
delay3[delay3 > 0] = normalized_positive_delay3.flatten()

# 仅选择大于零的元素
positive_snr3 = snr3[snr3 > 0].reshape(-1, 1)
# 使用 MinMaxScaler 进行归一化
scaler_snr3 = MinMaxScaler()
normalized_positive_snr3 = scaler_snr3.fit_transform(positive_snr3)
# 将归一化后的结果填充回原始向量中
snr3[snr3 > 0] = normalized_positive_snr3.flatten()

# 仅选择大于零的元素
positive_plr3 = plr3[plr3 > 0].reshape(-1, 1)
# 使用 MinMaxScaler 进行归一化
scaler_plr3 = MinMaxScaler()
normalized_positive_plr3 = scaler_plr3.fit_transform(positive_plr3)
# 将归一化后的结果填充回原始向量中
plr3[plr3 > 0] = normalized_positive_plr3.flatten()

# 仅选择大于零的元素
positive_plr0 = plr0[plr0 > 0].reshape(-1, 1)
# 使用 MinMaxScaler 进行归一化
scaler_plr0 = MinMaxScaler()
normalized_positive_plr0 = scaler_plr0.fit_transform(positive_plr0)
# 将归一化后的结果填充回原始向量中
plr0[plr0 > 0] = normalized_positive_plr0.flatten()


# 计算分割索引
split_index = int(0.7 * len(snr1))
# 将 snr 向量分成前70%和后30%
snr1_train = snr1[:split_index]
snr1_test = snr1[split_index:]
delay1_train = delay1[:split_index]
delay1_test = delay1[split_index:]
plr1_train = plr1[:split_index]
plr1_test = plr1[split_index:]


snr2_train = snr2[:split_index]
snr2_test = snr2[split_index:]
delay2_train = delay2[:split_index]
delay2_test = delay2[split_index:]
plr2_train = plr2[:split_index]
plr2_test = plr2[split_index:]


snr3_train = snr3[:split_index]
snr3_test = snr3[split_index:]
delay3_train = delay3[:split_index]
delay3_test = delay3[split_index:]
plr3_train = plr3[:split_index]
plr3_test = plr3[split_index:]


plr0_train = plr0[:split_index]
plr0_test = plr0[split_index:]


# 创建特征矩阵 X 和标签向量 y
X_train1 = np.column_stack((snr1_train, delay1_train))
Y_train1 = plr1_train
X_test1 = np.column_stack((snr1_test, delay1_test))
Y_test1 = plr1_test
Y_pre_pos1 = Y_test1[Y_test1 > 0].reshape(-1, 1)
Y_test1[Y_test1 > 0] = scaler_plr1.inverse_transform(Y_pre_pos1).flatten()


X_train2 = np.column_stack((snr2_train, delay2_train))
Y_train2 = plr2_train
X_test2 = np.column_stack((snr2_test, delay2_test))
Y_test2 = plr2_test
Y_pre_pos2 = Y_test2[Y_test2 > 0].reshape(-1, 1)
Y_test2[Y_test2 > 0] = scaler_plr2.inverse_transform(Y_pre_pos2).flatten()


X_train3 = np.column_stack((snr3_train, delay3_train))
Y_train3 = plr3_train
X_test3 = np.column_stack((snr3_test, delay3_test))
Y_test3 = plr3_test
Y_pre_pos3 = Y_test3[Y_test3 > 0].reshape(-1, 1)
Y_test3[Y_test3 > 0] = scaler_plr3.inverse_transform(Y_pre_pos3).flatten()

Y_train0 = plr0_train
Y_test0 = plr0_test
Y_pre_pos0 = Y_test0[Y_test0 > 0].reshape(-1, 1)
Y_test0[Y_test0 > 0] = scaler_plr0.inverse_transform(Y_pre_pos0).flatten()

reg_forest1 = RandomForestRegressor(n_estimators=5, max_depth=15, bootstrap=True)  # 使用100棵树，最大深度为30
reg_forest1.fit(X_train1, Y_train1)

# 进行预测
Y_T1 = reg_forest1.predict(X_train1)
Y_T1 = Y_T1.reshape(-1, 1)

reg_forest2 = RandomForestRegressor(n_estimators=5, max_depth=15, bootstrap=True)  # 使用100棵树，最大深度为30
reg_forest2.fit(X_train2, Y_train2)

# 进行预测
Y_T2 = reg_forest2.predict(X_train2)
Y_T2 = Y_T2.reshape(-1, 1)

reg_forest3 = RandomForestRegressor(n_estimators=5, max_depth=15, bootstrap=True)  # 使用100棵树，最大深度为30
reg_forest3.fit(X_train3, Y_train3)

# 进行预测
Y_T3 = reg_forest3.predict(X_train3)
Y_T3 = Y_T3.reshape(-1, 1)

reg_mlp1 = MLPRegressor(hidden_layer_sizes=(10, 10),  # 设置隐藏层神经元数量和层数
                        activation='relu',  # 使用ReLU作为激活函数
                        solver='adam',  # 使用Adam优化器
                        alpha=0.001,  # 正则化参数
                        batch_size='auto',  # 小批量梯度下降的批大小
                        learning_rate='constant',  # 使用固定的学习率
                        learning_rate_init=0.01,  # 初始学习率
                        max_iter=50,  # 最大迭代次数
                        tol=1e-4,  # 收敛容差
                        random_state=None)  # 随机种子

reg_mlp1.fit(X_train1, Y_train1)

# 进行预测
Y_F1 = reg_mlp1.predict(X_train1)
Y_F1 = Y_F1.reshape(-1, 1)


reg_mlp2 = MLPRegressor(hidden_layer_sizes=(10, 10),  # 设置隐藏层神经元数量和层数
                        activation='relu',  # 使用ReLU作为激活函数
                        solver='adam',  # 使用Adam优化器
                        alpha=0.001,  # 正则化参数
                        batch_size='auto',  # 小批量梯度下降的批大小
                        learning_rate='constant',  # 使用固定的学习率
                        learning_rate_init=0.01,  # 初始学习率
                        max_iter=50,  # 最大迭代次数
                        tol=1e-4,  # 收敛容差
                        random_state=None)  # 随机种子

reg_mlp2.fit(X_train2, Y_train2)

# 进行预测
Y_F2 = reg_mlp2.predict(X_train2)
Y_F2 = Y_F2.reshape(-1, 1)


reg_mlp3 = MLPRegressor(hidden_layer_sizes=(10, 10),  # 设置隐藏层神经元数量和层数
                        activation='relu',  # 使用ReLU作为激活函数
                        solver='adam',  # 使用Adam优化器
                        alpha=0.001,  # 正则化参数
                        batch_size='auto',  # 小批量梯度下降的批大小
                        learning_rate='constant',  # 使用固定的学习率
                        learning_rate_init=0.01,  # 初始学习率
                        max_iter=50,  # 最大迭代次数
                        tol=1e-4,  # 收敛容差
                        random_state=None)  # 随机种子

reg_mlp3.fit(X_train3, Y_train3)

# 进行预测
Y_F3 = reg_mlp3.predict(X_train3)
Y_F3 = Y_F3.reshape(-1, 1)

# 输入数据格式：
# input(seq_len, batch, input_size)
# h0(num_layers * num_directions, batch, hidden_size)
# c0(num_layers * num_directions, batch, hidden_size)
#
# 输出数据格式：
# output(seq_len, batch, hidden_size * num_directions)
# hn(num_layers * num_directions, batch, hidden_size)
# cn(num_layers * num_directions, batch, hidden_size)


# 定义 LSTM 模型
class LSTMModel(nn.Module):
    def __init__(self, input_size, hidden_size, output_size):
        super(LSTMModel, self).__init__()
        self.hidden_size = hidden_size
        self.lstm = nn.LSTM(input_size, hidden_size)
        self.fc = nn.Linear(hidden_size * 2, output_size)
        self.tanh = nn.ELU(alpha=50)

    def forward(self, x):
        # 前向传播
        out, _ = self.lstm(x)
        out = out.view(1, -1)
        out = self.fc(out)
        out = self.tanh(out)
        return out


# 创建模型实例
input_size = 1
hidden_size = 2
output_size = 1
Y_R1 = []

lstm1 = LSTMModel(input_size, hidden_size, output_size)
# 定义损失函数和优化器
criterion1 = nn.MSELoss()
optimizer1 = optim.Adam(lstm1.parameters(), lr=0.01)

# 准备训练数据

x_train_r1 = torch.tensor(X_train1, dtype=torch.float32).view(-1, 2)  # 输入维度为 (x, 2)
y_train_r1 = torch.tensor(Y_train1, dtype=torch.float32).view(-1, 1)            # 输出维度为 (x, 1)

# 训练模型
num_epochs = 5
for epoch in range(num_epochs):
    # 前向传播
    for batchnum in range(len(x_train_r1)):
        x_in = x_train_r1[batchnum]
        outputs = lstm1(x_in.view(len(x_in), 1, -1))  # 修改这里的输入维度
        if epoch == num_epochs-1:
            Y_R1.append(outputs.detach().numpy())
        # 计算损失
        loss = criterion1(outputs, y_train_r1[batchnum])

        # 反向传播和优化
        optimizer1.zero_grad()
        loss.backward()
        optimizer1.step()

    # # 打印训练信息
    # if (epoch+1) % 2 == 0:
    #     print(f'Epoch [{epoch+1}/{num_epochs}], Loss: {loss.item():.4f}')

Y_R1 = np.array(Y_R1).reshape(-1, 1)


Y_R2 = []

lstm2 = LSTMModel(input_size, hidden_size, output_size)
# 定义损失函数和优化器
criterion2 = nn.MSELoss()
optimizer2 = optim.Adam(lstm2.parameters(), lr=0.01)

# 准备训练数据

x_train_r2 = torch.tensor(X_train2, dtype=torch.float32).view(-1, 2)  # 输入维度为 (x, 2)
y_train_r2 = torch.tensor(Y_train2, dtype=torch.float32).view(-1, 1)            # 输出维度为 (x, 1)


for epoch in range(num_epochs):
    # 前向传播
    for batchnum in range(len(x_train_r2)):
        x_in = x_train_r2[batchnum]
        outputs = lstm2(x_in.view(len(x_in), 1, -1))  # 修改这里的输入维度
        if epoch == num_epochs-1:
            Y_R2.append(outputs.detach().numpy())
        # 计算损失
        loss = criterion2(outputs, y_train_r2[batchnum])

        # 反向传播和优化
        optimizer2.zero_grad()
        loss.backward()
        optimizer2.step()

    # # 打印训练信息
    # if (epoch+1) % 2 == 0:
    #     print(f'Epoch [{epoch+1}/{num_epochs}], Loss: {loss.item():.4f}')

Y_R2 = np.array(Y_R2).reshape(-1, 1)


Y_R3 = []

lstm3 = LSTMModel(input_size, hidden_size, output_size)
# 定义损失函数和优化器
criterion3 = nn.MSELoss()
optimizer3 = optim.Adam(lstm3.parameters(), lr=0.01)

# 准备训练数据

x_train_r3 = torch.tensor(X_train3, dtype=torch.float32).view(-1, 2)  # 输入维度为 (x, 2)
y_train_r3 = torch.tensor(Y_train3, dtype=torch.float32).view(-1, 1)            # 输出维度为 (x, 1)
y_train_r0 = torch.tensor(Y_train0, dtype=torch.float32).view(-1, 1)

for epoch in range(num_epochs):
    # 前向传播
    for batchnum in range(len(x_train_r3)):
        x_in = x_train_r3[batchnum]
        outputs = lstm3(x_in.view(len(x_in), 1, -1))  # 修改这里的输入维度
        if epoch == num_epochs-1:
            Y_R3.append(outputs.detach().numpy())
        # 计算损失
        loss = criterion3(outputs, y_train_r3[batchnum])

        # 反向传播和优化
        optimizer3.zero_grad()
        loss.backward()
        optimizer3.step()

    # # 打印训练信息
    # if (epoch+1) % 2 == 0:
    #     print(f'Epoch [{epoch+1}/{num_epochs}], Loss: {loss.item():.4f}')

Y_R3 = np.array(Y_R3).reshape(-1, 1)


class Combine(nn.Module):
    def __init__(self, input_size, trans_size, hidden_size):
        super(Combine, self).__init__()
        self.qmat = nn.Linear(input_size, trans_size)
        self.kmat = nn.Linear(input_size, trans_size)
        self.tanh = nn.ELU(alpha=50)

        self.hiddenlayer1 = nn.Linear(3, hidden_size)
        self.tanh1 = nn.ELU(alpha=50)

        self.hiddenlayer2 = nn.Linear(hidden_size, 1)
        self.tanh2 = nn.ELU(alpha=100)

    def forward(self, x1, x2, x3):
        # 前向传播
        q1 = self.qmat(x1)
        q2 = self.qmat(x2)
        q3 = self.qmat(x3)
        k1 = self.kmat(x1)
        k2 = self.kmat(x2)
        k3 = self.kmat(x3)

        beta11 = torch.sum(torch.mul(q1, k1))
        beta12 = torch.sum(torch.mul(q1, k2))
        beta13 = torch.sum(torch.mul(q1, k3))

        beta21 = torch.sum(torch.mul(q2, k1))
        beta22 = torch.sum(torch.mul(q2, k2))
        beta23 = torch.sum(torch.mul(q2, k3))

        beta31 = torch.sum(torch.mul(q3, k1))
        beta32 = torch.sum(torch.mul(q3, k2))
        beta33 = torch.sum(torch.mul(q3, k3))

        b1123 = F.softmax(torch.tensor([beta11.clone().detach(), beta12.clone().detach(), beta13.clone().detach()]), dim=0)
        b11, b12, b13 = b1123
        b2123 = F.softmax(torch.tensor([beta21.clone().detach(), beta22.clone().detach(), beta23.clone().detach()]), dim=0)
        b21, b22, b23 = b2123
        b3123 = F.softmax(torch.tensor([beta31.clone().detach(), beta32.clone().detach(), beta33.clone().detach()]), dim=0)
        b31, b32, b33 = b3123

        t1 = reg_forest1.predict(x1.view(1, -1))
        t1 = torch.tensor(t1).view(1, 1)
        f1 = reg_mlp1.predict(x1.view(1, -1))
        f1 = torch.tensor(f1).view(1, 1)
        r1 = lstm1(x1.view(len(x1), 1, -1))

        t2 = reg_forest2.predict(x2.view(1, -1))
        t2 = torch.tensor(t2).view(1, 1)
        f2 = reg_mlp2.predict(x2.view(1, -1))
        f2 = torch.tensor(f2).view(1, 1)
        r2 = lstm2(x2.view(len(x2), 1, -1))
        t3 = reg_forest3.predict(x3.view(1, -1))
        t3 = torch.tensor(t3).view(1, 1)
        f3 = reg_mlp3.predict(x3.view(1, -1))
        f3 = torch.tensor(f3).view(1, 1)
        r3 = lstm3(x3.view(len(x3), 1, -1))
        o1 = 1 / 3 * (t1 + f1 + r1)
        o2 = 1 / 3 * (t2 + f2 + r2)
        o3 = 1 / 3 * (t3 + f3 + r3)
        # o1 = r1
        # o2 = r2
        # o3 = r3
        # o1 = 1 / 2 * (t1 + r1)
        # o2 = 1 / 2 * (t2 + r2)
        # o3 = 1 / 2 * (t3 + r3)

        e1 = b11 * o1 + b12 * o1 + b13 * o1
        e2 = b21 * o2 + b22 * o2 + b23 * o2
        e3 = b31 * o3 + b32 * o3 + b33 * o3
        e1[e1 < -0.5] = -1.0
        e2[e2 < -0.3] = -1.0
        e3[e3 < -0.3] = -1.0
        e123 = torch.cat((e1, e2, e3), dim=1)
        e123 = e123.to(torch.float32)
        e = self.hiddenlayer1(e123)
        e = self.tanh1(e)
        e = self.hiddenlayer2(e)
        e0 = self.tanh2(e)
        return e1, e2, e3, e0


class Combineloss(nn.Module):
    def __init__(self):
        super(Combineloss, self).__init__()

    def forward(self, e1, e2, e3, e0, p1, p2, p3, p0):
        lam = 0.8
        closs = lam * (torch.pow((e1-p1), 2) + torch.pow((e2-p2), 2) + torch.pow((e3-p3), 2)) + (1 - lam) * torch.pow((e0-p0), 2)
        return closs

# input_size = 2
# trans_size = 3
# hidden_size = 3


com = Combine(input_size=2, hidden_size=3, trans_size=3)
# 定义损失函数和优化器
criterion_com = Combineloss()
optimizer_com = optim.Adam(com.parameters(), lr=0.01)

for epoch in range(8):
    # 前向传播
    for batchnum in range(len(x_train_r1)):
        x1 = x_train_r1[batchnum]
        x2 = x_train_r2[batchnum]
        x3 = x_train_r3[batchnum]
        y1 = y_train_r1[batchnum]
        y2 = y_train_r2[batchnum]
        y3 = y_train_r3[batchnum]
        y0 = y_train_r0[batchnum]

        e1, e2, e3, e0 = com(x1, x2, x3)
        # 计算损失
        loss_com = criterion_com(e1, e2, e3, e0, y1, y2, y3, y0)
        # 反向传播和优化
        optimizer_com.zero_grad()
        loss_com.backward()
        optimizer_com.step()

    # 打印训练信息
    if (epoch+1) % 1 == 0:
        print(f'Epoch [{epoch+1}/{num_epochs}], Loss: {loss_com.item():.4f}')

x_test_r1 = torch.tensor(X_test1, dtype=torch.float32).view(-1, 2)
y_test_r1 = torch.tensor(Y_test1, dtype=torch.float32).view(-1, 1)

x_test_r2 = torch.tensor(X_test2, dtype=torch.float32).view(-1, 2)
y_test_r2 = torch.tensor(Y_test2, dtype=torch.float32).view(-1, 1)

x_test_r3 = torch.tensor(X_test3, dtype=torch.float32).view(-1, 2)
y_test_r3 = torch.tensor(Y_test3, dtype=torch.float32).view(-1, 1)

y_test_r0 = torch.tensor(Y_test0, dtype=torch.float32).view(-1, 1)

y1_pre = []
y2_pre = []
y3_pre = []
y0_pre = []

for batchnum in range(len(x_test_r1)):
    x1 = x_test_r1[batchnum]
    x2 = x_test_r2[batchnum]
    x3 = x_test_r3[batchnum]
    y1 = y_test_r1[batchnum]
    y2 = y_test_r2[batchnum]
    y3 = y_test_r3[batchnum]
    y0 = y_test_r0[batchnum]
    e1, e2, e3, e0 = com(x1, x2, x3)
    if e0 > 0.5:
        e0 = e0 * 0.7
    y1_pre.append(e1)
    y2_pre.append(e2)
    y3_pre.append(e3)
    y0_pre.append(e0)


y1_pre = torch.cat(y1_pre, dim=0)
y2_pre = torch.cat(y2_pre, dim=0)
y3_pre = torch.cat(y3_pre, dim=0)
y0_pre = torch.cat(y0_pre, dim=0)

y1_pre = y1_pre.detach().numpy().reshape(-1, 1)
y2_pre = y2_pre.detach().numpy().reshape(-1, 1)
y3_pre = y3_pre.detach().numpy().reshape(-1, 1)
y0_pre = y0_pre.detach().numpy().reshape(-1, 1)

Y_pre_pos1 = y1_pre[y1_pre > 0].reshape(-1, 1)
y1_pre[y1_pre > 0] = scaler_plr1.inverse_transform(Y_pre_pos1).flatten()

Y_pre_pos2 = y2_pre[y2_pre > 0].reshape(-1, 1)
y2_pre[y2_pre > 0] = scaler_plr2.inverse_transform(Y_pre_pos2).flatten()

Y_pre_pos3 = y3_pre[y3_pre > 0].reshape(-1, 1)
y3_pre[y3_pre > 0] = scaler_plr3.inverse_transform(Y_pre_pos3).flatten()


# 计算均方误差（MSE）
mse0 = mean_squared_error(y0_pre, y_test_r0)
# 计算平均绝对误差（MAE）
mae0 = mean_absolute_error(y0_pre, y_test_r0)
# 计算均方根误差（RMSE）
rmse0 = np.sqrt(mse0)
# 计算平均平方对数误差（MSLE）

# 计算 R 平方值（R2）
r20 = r2_score(y0_pre, y_test_r0)

mse1 = mean_squared_error(y1_pre, y_test_r1)
# 计算平均绝对误差（MAE）
mae1 = mean_absolute_error(y1_pre, y_test_r1)
# 计算均方根误差（RMSE）
rmse1 = np.sqrt(mse1)
# 计算平均平方对数误差（MSLE）

# 计算 R 平方值（R2）
r21 = r2_score(y1_pre, y_test_r1)

mse2 = mean_squared_error(y2_pre, y_test_r2)
# 计算平均绝对误差（MAE）
mae2 = mean_absolute_error(y2_pre, y_test_r2)
# 计算均方根误差（RMSE）
rmse2 = np.sqrt(mse2)
# 计算平均平方对数误差（MSLE）

# 计算 R 平方值（R2）
r22 = r2_score(y2_pre, y_test_r2)

mse3 = mean_squared_error(y3_pre, y_test_r3)
# 计算平均绝对误差（MAE）
mae3 = mean_absolute_error(y3_pre, y_test_r3)
# 计算均方根误差（RMSE）
rmse3 = np.sqrt(mse3)
# 计算平均平方对数误差（MSLE）

# 计算 R 平方值（R2）
r23 = r2_score(y3_pre, y_test_r3)
# Y_R1 = np.array(Y_R1).reshape(-1, 1)
#
#
#
# o1 = 1/3*(Y_R1 + Y_T1 + Y_F1)
# o2 = 1/3*(Y_R2 + Y_T2 + Y_F2)
# o3 = 1/3*(Y_R3 + Y_T3 + Y_F3)

# # 测试模型
# x_test = torch.tensor([[1.0, 2.0]], dtype=torch.float32)
# with torch.no_grad():
#     pred = model(x_test.view(1, 1, -1))  # 修改这里的输入维度
#     print("测试结果:", pred.item())


# np.save('y0_pre5.npy', y0_pre)
# np.save('y1_pre5.npy', y1_pre)
# np.save('y2_pre5.npy', y2_pre)
# np.save('y3_pre5.npy', y3_pre)
#
# np.save('y0.npy', Y_test0)
# np.save('y1.npy', Y_test1)
# np.save('y2.npy', Y_test2)
# np.save('y3.npy', Y_test3)


print(1)

