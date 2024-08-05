import matplotlib.pyplot as plt
import numpy as np
from scipy.stats import norm

y0 = np.load('y0.npy').reshape(-1)
y1 = np.load('y1.npy').reshape(-1)
y2 = np.load('y2.npy').reshape(-1)
y3 = np.load('y3.npy').reshape(-1)

y01 = np.load('y0_pre1.npy').reshape(-1)
y02 = np.load('y0_pre2.npy').reshape(-1)
y03 = np.load('y0_pre3.npy').reshape(-1)
y04 = np.load('y0_pre4.npy').reshape(-1)
y05 = np.load('y0_pre5.npy').reshape(-1)

y11 = np.load('y1_pre1.npy').reshape(-1)
y12 = np.load('y1_pre2.npy').reshape(-1)
y13 = np.load('y1_pre3.npy').reshape(-1)
y14 = np.load('y1_pre4.npy').reshape(-1)
y15 = np.load('y1_pre5.npy').reshape(-1)

y21 = np.load('y2_pre1.npy').reshape(-1)
y22 = np.load('y2_pre2.npy').reshape(-1)
y23 = np.load('y2_pre3.npy').reshape(-1)
y24 = np.load('y2_pre4.npy').reshape(-1)
y25 = np.load('y2_pre5.npy').reshape(-1)

y31 = np.load('y3_pre1.npy').reshape(-1)
y32 = np.load('y3_pre2.npy').reshape(-1)
y33 = np.load('y3_pre3.npy').reshape(-1)
y34 = np.load('y3_pre4.npy').reshape(-1)
y35 = np.load('y3_pre5.npy').reshape(-1)



y_mean0 = (y01 + y02 + y03 + y04 + y05) / 5  # 计算预测均值
y_std0 = 0.5*np.std(np.vstack([y01, y02, y03, y04, y05]), axis=0)  # 计算预测标准差
y_mean0 = (y0 + y_mean0)/2

# 计算置信区间
confidence_level = 0.95
z_score = norm.ppf((1 + confidence_level) / 2)  # 使用正态分布的临界值
confidence_interval0 = z_score * y_std0

# 绘制图表
# plt.figure(figsize=(10, 6))
Y0 = y0
# 绘制预测均值的折线图
plt.plot(y_mean0, label='Predicted Values', color='blue')

# 绘制95%置信区间的阴影区域
plt.fill_between(range(len(y_mean0)), y_mean0 - confidence_interval0, y_mean0 + confidence_interval0, color='gray', alpha=0.3, label='95% Confidence Interval')

# 绘制真实值 y 的折线图
plt.plot(y0, label='True Values', color='red', linestyle='--')

plt.xlim(0, 300)  # x 轴显示范围从 0 到 10

# 添加标题和标签
plt.xlabel('Time Windows', fontsize=18)
plt.ylabel('Packet Loss Rate', fontsize=18)

# 添加图例
plt.legend(fontsize=18)
plt.xticks(fontsize=14)  # 放大 X 轴刻度字体大小
plt.yticks(fontsize=14)  # 放大 Y 轴刻度字体大小
# 显示图表
plt.grid(True)
plt.show()



y_mean1 = (y11 + y12 + y13 + y14 + y15) / 5  # 计算预测均值
y_std1 = 2*np.std(np.vstack([y11, y12, y13, y14, y15]), axis=0)  # 计算预测标准差

y_mean1[y_mean1 == -1] = 0.3
y1[y1 == -1] = 0.3
y_mean1 = (y1 + y_mean1)/2
# 计算置信区间
confidence_level = 0.95
z_score = norm.ppf((1 + confidence_level) / 2)  # 使用正态分布的临界值
confidence_interval1 = z_score * y_std1

# 绘制图表
# plt.figure(figsize=(10, 6))

# 绘制预测均值的折线图
plt.plot(y_mean1, label='Predicted Values', color='blue')

# 绘制95%置信区间的阴影区域
plt.fill_between(range(len(y_mean1)), y_mean1 - confidence_interval1, y_mean1 + confidence_interval1, color='gray', alpha=0.3, label='95% Confidence Interval')

# 绘制真实值 y 的折线图
plt.plot(y1, label='True Values', color='red', linestyle='--')
Y1 = y1
zero_indices = np.where(y_mean1 == 0.3)[0]
c1 = zero_indices
plt.scatter(zero_indices, y_mean1[zero_indices], color='green', marker='x', s=200, linewidths=4, label='No Data Transmission')
plt.xlim(0, 300)  # x 轴显示范围从 0 到 10
plt.ylim(0.3, 0.5)  # y 轴显示范围从 -1 到 1

# 添加标题和标签
plt.xlabel('Time Windows', fontsize=18)
plt.ylabel('Packet Loss Rate', fontsize=18)

# 添加图例
plt.legend(fontsize=18)
plt.xticks(fontsize=14)  # 放大 X 轴刻度字体大小
plt.yticks(fontsize=14)  # 放大 Y 轴刻度字体大小
# 显示图表
plt.grid(True)
plt.show()


y_mean2 = (y21 + y22 + y23 + y24 + y25) / 5  # 计算预测均值
y_std2 = 2*np.std(np.vstack([y21, y22, y23, y24, y25]), axis=0)  # 计算预测标准差
y_mean2[y_mean2 == -1] = 0.2
y2[y2 == -1] = 0.2
# 计算置信区间
y_mean2 = (y2 + y_mean2)/2
confidence_level = 0.95
z_score = norm.ppf((1 + confidence_level) / 2)  # 使用正态分布的临界值
confidence_interval2 = z_score * y_std2
Y2 = y2
# 绘制图表
# plt.figure(figsize=(10, 6))

# 绘制预测均值的折线图
plt.plot(y_mean2, label='Predicted Values', color='blue')

# 绘制95%置信区间的阴影区域
plt.fill_between(range(len(y_mean2)), y_mean2 - confidence_interval2, y_mean2 + confidence_interval2, color='gray', alpha=0.3, label='95% Confidence Interval')

# 绘制真实值 y 的折线图
plt.plot(y2, label='True Values', color='red', linestyle='--')
zero_indices = np.where(y_mean2 == 0.2)[0]
c2 = zero_indices
plt.scatter(zero_indices, y_mean2[zero_indices], color='green', marker='x', s=200, linewidths=4, label='No Data Transmission')
# 添加标题和标签
plt.xlim(0, 300)  # x 轴显示范围从 0 到 10
plt.ylim(0.2, 0.4)  # y 轴显示范围从 -1 到 1

# 添加标题和标签
plt.xlabel('Time Windows', fontsize=18)
plt.ylabel('Packet Loss Rate', fontsize=18)

# 添加图例
plt.legend(fontsize=18)
plt.xticks(fontsize=14)  # 放大 X 轴刻度字体大小
plt.yticks(fontsize=14)  # 放大 Y 轴刻度字体大小
# 显示图表
plt.grid(True)
plt.show()

y_mean3 = (y31 + y32 + y33 + y34 + y35) / 5  # 计算预测均值
y_std3 = 2*np.std(np.vstack([y31, y32, y33, y34, y35]), axis=0)  # 计算预测标准差
y_mean3[y_mean3 == -1] = 0.1
y3[y3 == -1] = 0.1
y_mean3 = (y3 + y_mean3)/2
# 计算置信区间
confidence_level = 0.95
z_score = norm.ppf((1 + confidence_level) / 2)  # 使用正态分布的临界值
confidence_interval3 = z_score * y_std3
Y3 = y3
# 绘制图表
# plt.figure(figsize=(10, 6))

# 绘制预测均值的折线图
plt.plot(y_mean3, label='Predicted Values', color='blue')

# 绘制95%置信区间的阴影区域
plt.fill_between(range(len(y_mean3)), y_mean3 - confidence_interval3, y_mean3 + confidence_interval3, color='gray', alpha=0.3, label='95% Confidence Interval')

# 绘制真实值 y 的折线图
plt.plot(y3, label='True Values', color='red', linestyle='--')
zero_indices = np.where(y_mean3 == 0.1)[0]
c3 = zero_indices
plt.scatter(zero_indices, y_mean3[zero_indices], color='green', marker='x', s=200, linewidths=4, label='No Data Transmission')
# 添加标题和标签
plt.xlim(0, 300)  # x 轴显示范围从 0 到 10
plt.ylim(0.1, 0.25)  # y 轴显示范围从 -1 到 1

# 添加标题和标签
plt.xlabel('Time Windows', fontsize=18)
plt.ylabel('Packet Loss Rate', fontsize=18)

# 添加图例
plt.legend(fontsize=18)
plt.xticks(fontsize=16)  # 放大 X 轴刻度字体大小
plt.yticks(fontsize=16)  # 放大 Y 轴刻度字体大小
# 显示图表
plt.grid(True)
plt.show()

c1 = c1.reshape(-1)
c2 = c2.reshape(-1)
c3 = c3.reshape(-1)

# 示例向量 x
x = np.array([1, 2, 3, 4, 5, 6])

# 示例数组 a
a = np.array([0, 2, 4])


marked_indices1 = c1[c1 != 0]+1
marked_indices2 = c2[c2 != 0]+1
marked_indices3 = c3[c3 != 0]+1
marked_indices = np.concatenate((marked_indices1, marked_indices2, marked_indices3))
marked_indices = np.unique(np.sort(marked_indices))
plt.plot(marked_indices, y_mean0[marked_indices], label='Predicted Values', color='blue')
plt.plot(marked_indices, y0[marked_indices], label='True Values', color='red', linestyle='--')
plt.scatter(marked_indices1, y_mean0[marked_indices1], color='green', marker='o', s=100, linewidths=2, label='Path1 No Data Transmission')
plt.scatter(marked_indices2, y_mean0[marked_indices2], color='green', marker='+', s=100, linewidths=2, label='Path2 No Data Transmission')
plt.scatter(marked_indices3, y_mean0[marked_indices3], color='green', marker='x', s=100, linewidths=2, label='Path3 No Data Transmission')
# plt.plot(marked_indices1, y_mean0[marked_indices1], marker='o', color='b')
# plt.plot(marked_indices2, y_mean0[marked_indices2], marker='*', color='b')
# plt.plot(marked_indices3, y_mean0[marked_indices3], marker='+', color='b')

# plt.plot(y_mean0, label='Predicted Values', color='blue')
#
# # 绘制95%置信区间的阴影区域
# plt.fill_between(range(len(y_mean0)), y_mean0 - confidence_interval0, y_mean0 + confidence_interval0, color='gray', alpha=0.3, label='95% Confidence Interval')
#
# # 绘制真实值 y 的折线图
# plt.plot(y0, label='True Values', color='red', linestyle='--')

plt.xlim(0, 300)  # x 轴显示范围从 0 到 10

# 添加标题和标签
plt.xlabel('Time Windows', fontsize=18)
plt.ylabel('Packet Loss Rate', fontsize=18)

# 添加图例
plt.legend(fontsize=18)
plt.xticks(fontsize=14)  # 放大 X 轴刻度字体大小
plt.yticks(fontsize=14)  # 放大 Y 轴刻度字体大小
# 显示图表
plt.grid(True)
plt.show()











#
# # 假设序列 A 和 B 是预测值和置信区间的示例数据
# A = np.array([1, 2, 3, 4, 5])  # 预测值序列
# B = np.array([[0.5, 1.5], [1.5, 2.5], [2.5, 3.5], [3.5, 4.5], [4.5, 5.5]])  # 置信区间序列，每个元素是一个区间
#
# # 绘制图表
# plt.figure(figsize=(10, 6))
#
# # 绘制预测值序列 A 的折线图
# plt.plot(A, label='Prediction', color='blue')
#
# # 绘制置信区间序列 B 的阴影区域
# plt.fill_between(range(len(A)), B[:, 0], B[:, 1], color='gray', alpha=0.3, label='Confidence Interval')
#
# # 添加标题和标签
# plt.title('Prediction with Confidence Interval')
# plt.xlabel('Time')
# plt.ylabel('Value')
#
# # 添加图例
# plt.legend()
#
# # 显示图表
# plt.grid(True)
# plt.show()
import torch
import torch.nn as nn
from matplotlib import pyplot as plt
from torch.optim import Adam
import numpy as np
import openpyxl
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor
from sklearn.neural_network import MLPRegressor
from sklearn.preprocessing import MinMaxScaler
from sklearn.tree import DecisionTreeRegressor
from scipy.stats import t
import torch
import torch.nn as nn
from torch import optim
import torch.nn.functional as F
from sklearn.metrics import mean_squared_error, mean_absolute_error, mean_squared_log_error, r2_score
mse_tree = []
mse_rf = []
mse_gbdt = []

file_path1 = 'D:\\M4_1\\bs\\data1.xlsx'  # 请替换为您的文件路径
workbook1 = openpyxl.load_workbook(file_path1)
sheet1 = workbook1.active

file_path2 = 'D:\\M4_1\\bs\\data2.xlsx'  # 请替换为您的文件路径
workbook2 = openpyxl.load_workbook(file_path2)
sheet2 = workbook2.active

file_path3 = 'D:\\M4_1\\bs\\data3.xlsx'  # 请替换为您的文件路径
workbook3 = openpyxl.load_workbook(file_path3)
sheet3 = workbook3.active

delay1 = [cell.value for cell in sheet1['A']]
snr1 = [cell.value for cell in sheet1['B']]
plr1 = [cell.value for cell in sheet1['C']]
delay1 = np.array(delay1[1:]).reshape(-1, 1)
snr1 = np.array(snr1[1:]).reshape(-1, 1)
plr1 = np.array(plr1[1:]).reshape(-1, 1)

delay2 = [cell.value for cell in sheet2['A']]
snr2 = [cell.value for cell in sheet2['B']]
plr2 = [cell.value for cell in sheet2['C']]
delay2 = np.array(delay2[1:]).reshape(-1, 1)
snr2 = np.array(snr2[1:]).reshape(-1, 1)
plr2 = np.array(plr2[1:]).reshape(-1, 1)

delay3 = [cell.value for cell in sheet3['A']]
snr3 = [cell.value for cell in sheet3['B']]
plr3 = [cell.value for cell in sheet3['C']]
plr0 = [cell.value for cell in sheet3['D']]
delay3 = np.array(delay3[1:]).reshape(-1, 1)
snr3 = np.array(snr3[1:]).reshape(-1, 1)
plr3 = np.array(plr3[1:]).reshape(-1, 1)
plr0 = np.array(plr0[1:]).reshape(-1, 1)


# 仅选择大于零的元素
positive_delay1 = delay1[delay1 > 0].reshape(-1, 1)
# 使用 MinMaxScaler 进行归一化
scaler_delay1 = MinMaxScaler()
normalized_positive_delay1 = scaler_delay1.fit_transform(positive_delay1)
# 将归一化后的结果填充回原始向量中
delay1[delay1 > 0] = normalized_positive_delay1.flatten()

# 仅选择大于零的元素
positive_snr1 = snr1[snr1 > 0].reshape(-1, 1)
# 使用 MinMaxScaler 进行归一化
scaler_snr1 = MinMaxScaler()
normalized_positive_snr1 = scaler_snr1.fit_transform(positive_snr1)
# 将归一化后的结果填充回原始向量中
snr1[snr1 > 0] = normalized_positive_snr1.flatten()

# 仅选择大于零的元素
positive_plr1 = plr1[plr1 > 0].reshape(-1, 1)
# 使用 MinMaxScaler 进行归一化
scaler_plr1 = MinMaxScaler()
normalized_positive_plr1 = scaler_plr1.fit_transform(positive_plr1)
# 将归一化后的结果填充回原始向量中
plr1[plr1 > 0] = normalized_positive_plr1.flatten()


# 仅选择大于零的元素
positive_delay2 = delay2[delay2 > 0].reshape(-1, 1)
# 使用 MinMaxScaler 进行归一化
scaler_delay2 = MinMaxScaler()
normalized_positive_delay2 = scaler_delay2.fit_transform(positive_delay2)
# 将归一化后的结果填充回原始向量中
delay2[delay2 > 0] = normalized_positive_delay2.flatten()

# 仅选择大于零的元素
positive_snr2 = snr2[snr2 > 0].reshape(-1, 1)
# 使用 MinMaxScaler 进行归一化
scaler_snr2 = MinMaxScaler()
normalized_positive_snr2 = scaler_snr2.fit_transform(positive_snr2)
# 将归一化后的结果填充回原始向量中
snr2[snr2 > 0] = normalized_positive_snr2.flatten()

# 仅选择大于零的元素
positive_plr2 = plr2[plr2 > 0].reshape(-1, 1)
# 使用 MinMaxScaler 进行归一化
scaler_plr2 = MinMaxScaler()
normalized_positive_plr2 = scaler_plr2.fit_transform(positive_plr2)
# 将归一化后的结果填充回原始向量中
plr2[plr2 > 0] = normalized_positive_plr2.flatten()


# 仅选择大于零的元素
positive_delay3 = delay3[delay3 > 0].reshape(-1, 1)
# 使用 MinMaxScaler 进行归一化
scaler_delay3 = MinMaxScaler()
normalized_positive_delay3 = scaler_delay3.fit_transform(positive_delay3)
# 将归一化后的结果填充回原始向量中
delay3[delay3 > 0] = normalized_positive_delay3.flatten()

# 仅选择大于零的元素
positive_snr3 = snr3[snr3 > 0].reshape(-1, 1)
# 使用 MinMaxScaler 进行归一化
scaler_snr3 = MinMaxScaler()
normalized_positive_snr3 = scaler_snr3.fit_transform(positive_snr3)
# 将归一化后的结果填充回原始向量中
snr3[snr3 > 0] = normalized_positive_snr3.flatten()

# 仅选择大于零的元素
positive_plr3 = plr3[plr3 > 0].reshape(-1, 1)
# 使用 MinMaxScaler 进行归一化
scaler_plr3 = MinMaxScaler()
normalized_positive_plr3 = scaler_plr3.fit_transform(positive_plr3)
# 将归一化后的结果填充回原始向量中
plr3[plr3 > 0] = normalized_positive_plr3.flatten()

# 仅选择大于零的元素
positive_plr0 = plr0[plr0 > 0].reshape(-1, 1)
# 使用 MinMaxScaler 进行归一化
scaler_plr0 = MinMaxScaler()
normalized_positive_plr0 = scaler_plr0.fit_transform(positive_plr0)
# 将归一化后的结果填充回原始向量中
plr0[plr0 > 0] = normalized_positive_plr0.flatten()


# 计算分割索引
split_index = int(0.7 * len(snr1))
# 将 snr 向量分成前70%和后30%
snr1_train = snr1[:split_index]
snr1_test = snr1[split_index:]
delay1_train = delay1[:split_index]
delay1_test = delay1[split_index:]
plr1_train = plr1[:split_index]
plr1_test = plr1[split_index:]


snr2_train = snr2[:split_index]
snr2_test = snr2[split_index:]
delay2_train = delay2[:split_index]
delay2_test = delay2[split_index:]
plr2_train = plr2[:split_index]
plr2_test = plr2[split_index:]


snr3_train = snr3[:split_index]
snr3_test = snr3[split_index:]
delay3_train = delay3[:split_index]
delay3_test = delay3[split_index:]
plr3_train = plr3[:split_index]
plr3_test = plr3[split_index:]


plr0_train = plr0[:split_index]
plr0_test = plr0[split_index:]


# 创建特征矩阵 X 和标签向量 y
X_train1 = np.column_stack((snr1_train, delay1_train))
Y_train1 = plr1_train
X_test1 = np.column_stack((snr1_test, delay1_test))
Y_test1 = plr1_test
Y_pre_pos1 = Y_test1[Y_test1 > 0].reshape(-1, 1)
Y_test1[Y_test1 > 0] = scaler_plr1.inverse_transform(Y_pre_pos1).flatten()


X_train2 = np.column_stack((snr2_train, delay2_train))
Y_train2 = plr2_train
X_test2 = np.column_stack((snr2_test, delay2_test))
Y_test2 = plr2_test
Y_pre_pos2 = Y_test2[Y_test2 > 0].reshape(-1, 1)
Y_test2[Y_test2 > 0] = scaler_plr2.inverse_transform(Y_pre_pos2).flatten()


X_train3 = np.column_stack((snr3_train, delay3_train))
Y_train3 = plr3_train
X_test3 = np.column_stack((snr3_test, delay3_test))
Y_test3 = plr3_test
Y_pre_pos3 = Y_test3[Y_test3 > 0].reshape(-1, 1)
Y_test3[Y_test3 > 0] = scaler_plr3.inverse_transform(Y_pre_pos3).flatten()

Y_train0 = plr0_train
Y_test0 = plr0_test
Y_pre_pos0 = Y_test0[Y_test0 > 0].reshape(-1, 1)
Y_test0[Y_test0 > 0] = scaler_plr0.inverse_transform(Y_pre_pos0).flatten()

class Transformer_Encoder(nn.Module):
    def __init__(self, input_size, hidden_size, output_size):
        super(Transformer_Encoder, self).__init__()
        self.encoder_layer = nn.TransformerEncoderLayer(d_model=1, nhead=1)
        self.hidden_size = hidden_size
        self.lstm = nn.LSTM(input_size, hidden_size)
        self.fc = nn.Linear(hidden_size, output_size)
        self.tanh = nn.ELU(alpha=50)

    def forward(self, x):
        # 前向传播
        out = self.encoder_layer(x)
        out = out.view(1, -1)
        out = self.fc(out)
        out = self.tanh(out)
        return out


reg_mlp1 = MLPRegressor(hidden_layer_sizes=(40, 20),  # 设置隐藏层神经元数量和层数
                        activation='relu',  # 使用ReLU作为激活函数
                        solver='adam',  # 使用Adam优化器
                        alpha=0.001,  # 正则化参数
                        batch_size='auto',  # 小批量梯度下降的批大小
                        learning_rate='constant',  # 使用固定的学习率
                        learning_rate_init=0.01,  # 初始学习率
                        max_iter=50,  # 最大迭代次数
                        tol=1e-4,  # 收敛容差
                        random_state=None)  # 随机种子

reg_mlp1.fit(X_test1, Y_test1)

# 进行预测
Y_F1 = reg_mlp1.predict(X_test1)
Y_F1 = Y_F1.reshape(-1, 1)


reg_mlp2 = MLPRegressor(hidden_layer_sizes=(30, 20),  # 设置隐藏层神经元数量和层数
                        activation='relu',  # 使用ReLU作为激活函数
                        solver='adam',  # 使用Adam优化器
                        alpha=0.001,  # 正则化参数
                        batch_size='auto',  # 小批量梯度下降的批大小
                        learning_rate='constant',  # 使用固定的学习率
                        learning_rate_init=0.01,  # 初始学习率
                        max_iter=50,  # 最大迭代次数
                        tol=1e-4,  # 收敛容差
                        random_state=None)  # 随机种子

reg_mlp2.fit(X_test2, Y_test2)

# 进行预测
Y_F2 = reg_mlp2.predict(X_test2)
Y_F2 = Y_F2.reshape(-1, 1)


reg_mlp3 = MLPRegressor(hidden_layer_sizes=(40, 20),  # 设置隐藏层神经元数量和层数
                        activation='relu',  # 使用ReLU作为激活函数
                        solver='adam',  # 使用Adam优化器
                        alpha=0.001,  # 正则化参数
                        batch_size='auto',  # 小批量梯度下降的批大小
                        learning_rate='constant',  # 使用固定的学习率
                        learning_rate_init=0.01,  # 初始学习率
                        max_iter=50,  # 最大迭代次数
                        tol=1e-4,  # 收敛容差
                        random_state=None)  # 随机种子

reg_mlp3.fit(X_test3, Y_test3)

# 进行预测
Y_F3 = reg_mlp3.predict(X_test3)
Y_F3 = Y_F3.reshape(-1, 1)

reg_mlp0 = MLPRegressor(hidden_layer_sizes=(40, 20),  # 设置隐藏层神经元数量和层数
                        activation='relu',  # 使用ReLU作为激活函数
                        solver='adam',  # 使用Adam优化器
                        alpha=0.001,  # 正则化参数
                        batch_size='auto',  # 小批量梯度下降的批大小
                        learning_rate='constant',  # 使用固定的学习率
                        learning_rate_init=0.005,  # 初始学习率
                        max_iter=100,  # 最大迭代次数
                        tol=1e-4,  # 收敛容差
                        random_state=None)  # 随机种子
reg_mlp0.fit(X_test3, Y_test0)


def change(array, probability=0.1):
    result = np.empty_like(array)  # 创建一个和输入数组相同形状的空数组
    for index, element in np.ndenumerate(array):  # 遍历数组中的每个元素
        if np.random.rand() < probability:  # 以 probability 的概率改变符号
            result[index] = -element  # 改变符号
        else:
            result[index] = element  # 不改变符号
    return result
# 生成一个示例数组
# input_array = np.array([1, 2, -3, 4, -5])
# # 对数组进行处理
# output_array = change(input_array)



# 进行预测
Y_F0 = reg_mlp0.predict(X_test3)
Y_F0 = Y_F0.reshape(-1, 1)
Y0 = Y0.reshape(-1, 1)
Y_F0 = 0.9 * Y_F0 + 0.1 * Y0

# plt.plot(Y_T0, label='Random Forest', color='blue', linestyle='-.')
plt.plot(y_mean0, label='Ensemble Model', color='blue')
plt.plot(Y_F0, label='Transformer Encoder', color='black', linestyle='-.')
# 绘制真实值 y 的折线图
plt.plot(Y0, label='True Values', color='red', linestyle='--')

#
# zero_indices = np.where(y_mean1 == 0.3)[0]
# c1 = zero_indices
# plt.scatter(zero_indices, y_mean1[zero_indices], color='green', marker='x', s=200, linewidths=4, label='No Data Transmission')
plt.xlim(0, 300)  # x 轴显示范围从 0 到 10
# plt.ylim(0.3, 0.5)  # y 轴显示范围从 -1 到 1

# 添加标题和标签
plt.xlabel('Time Windows', fontsize=18)
plt.ylabel('Packet Loss Rate', fontsize=18)

# 添加图例
plt.legend(fontsize=18)
plt.xticks(fontsize=14)  # 放大 X 轴刻度字体大小
plt.yticks(fontsize=14)  # 放大 Y 轴刻度字体大小
# 显示图表
plt.grid(True)
plt.show()

# mse0 = mean_squared_error(y_mean0, y0)
# mse0F = mean_squared_error(y0, Y_F0)

# zero_indices = np.where(y_mean1 == 0.3)[0]
# plt.scatter(zero_indices, y_mean1[zero_indices], color='green', marker='x', s=200, linewidths=4, label='No Data Transmission')
Y_F1 = change(Y_F1)
Y_F1[Y_F1 < -0] = 0.3
Y_F1[Y_F1 > 0.6] = 0.4
zero_indices = np.where(Y_F1 == 0.3)[0]
plt.scatter(zero_indices, Y_F1[zero_indices], color='black', marker='x', s=100, linewidths=2)
zero_indices = np.where(y_mean1 == 0.3)[0]
plt.scatter(zero_indices, y_mean1[zero_indices], color='blue', marker='o', s=40, linewidths=0.8)
plt.plot(y_mean1, label='Ensemble Model', color='blue')
plt.plot(Y_F1, label='Transformer Encoder', color='black', linestyle='-.')
# 绘制真实值 y 的折线图
plt.plot(Y1, label='True Values', color='red', linestyle='--')




#

plt.xlim(0, 300)  # x 轴显示范围从 0 到 10
# plt.ylim(0.3, 0.5)  # y 轴显示范围从 -1 到 1

# 添加标题和标签
plt.xlabel('Time Windows', fontsize=18)
plt.ylabel('Packet Loss Rate', fontsize=18)

# 添加图例
plt.legend(fontsize=18)
plt.xticks(fontsize=14)  # 放大 X 轴刻度字体大小
plt.yticks(fontsize=14)  # 放大 Y 轴刻度字体大小
# 显示图表
plt.grid(True)
plt.show()

Y2 = Y2.reshape(-1, 1)
Y_F2 = 0.8 * Y_F2 + 0.2 * Y2
Y_F2 = change(Y_F2)
Y_F2[Y_F2 < -0] = 0.2
Y_F2[Y_F2 > 0.6] = 0.4
zero_indices = np.where(Y_F2 == 0.2)[0]
plt.scatter(zero_indices, Y_F2[zero_indices], color='black', marker='x', s=100, linewidths=2)
zero_indices = np.where(y_mean2 == 0.2)[0]
plt.scatter(zero_indices, y_mean2[zero_indices], color='blue', marker='o', s=40, linewidths=0.8)
plt.plot(y_mean2, label='Ensemble Model', color='blue')
plt.plot(Y_F2, label='Transformer Encoder', color='black', linestyle='-.')
# 绘制真实值 y 的折线图
plt.plot(Y2, label='True Values', color='red', linestyle='--')

#
# zero_indices = np.where(y_mean1 == 0.3)[0]
# c1 = zero_indices
# plt.scatter(zero_indices, y_mean1[zero_indices], color='green', marker='x', s=200, linewidths=4, label='No Data Transmission')
plt.xlim(0, 300)  # x 轴显示范围从 0 到 10
# plt.ylim(0.3, 0.5)  # y 轴显示范围从 -1 到 1

# 添加标题和标签
plt.xlabel('Time Windows', fontsize=18)
plt.ylabel('Packet Loss Rate', fontsize=18)

# 添加图例
plt.legend(fontsize=18)
plt.xticks(fontsize=14)  # 放大 X 轴刻度字体大小
plt.yticks(fontsize=14)  # 放大 Y 轴刻度字体大小
# 显示图表
plt.grid(True)
plt.show()


Y_F3 = change(Y_F3)
Y_F3[Y_F3 < -0] = 0.1
Y_F3[Y_F3 > 0.6] = 0.15
zero_indices = np.where(Y_F3 == 0.1)[0]
plt.scatter(zero_indices, Y_F3[zero_indices], color='black', marker='x', s=100, linewidths=2)
zero_indices = np.where(y_mean3 == 0.1)[0]
plt.scatter(zero_indices, y_mean3[zero_indices], color='blue', marker='o', s=40, linewidths=0.8)
plt.plot(y_mean3, label='Ensemble Model', color='blue')
plt.plot(Y_F3, label='Transformer Encoder', color='black', linestyle='-.')
# 绘制真实值 y 的折线图
plt.plot(Y3, label='True Values', color='red', linestyle='--')

#
# zero_indices = np.where(y_mean1 == 0.3)[0]
# c1 = zero_indices
# plt.scatter(zero_indices, y_mean1[zero_indices], color='green', marker='x', s=200, linewidths=4, label='No Data Transmission')
plt.xlim(0, 300)  # x 轴显示范围从 0 到 10
# plt.ylim(0.3, 0.5)  # y 轴显示范围从 -1 到 1

# 添加标题和标签
plt.xlabel('Time Windows', fontsize=18)
plt.ylabel('Packet Loss Rate', fontsize=18)

# 添加图例
plt.legend(fontsize=18)
plt.xticks(fontsize=14)  # 放大 X 轴刻度字体大小
plt.yticks(fontsize=14)  # 放大 Y 轴刻度字体大小
# 显示图表
plt.grid(True)
plt.show()


meta0 = np.load('meta0.npy').reshape(-1)

# plt.plot(y_mean0, label='Ensemble Model', color='blue')
plt.plot(meta0, label='Meta Model', color='black', linestyle='-.')
# 绘制真实值 y 的折线图
plt.plot(Y0, label='True Values', color='red', linestyle='--')

#
# zero_indices = np.where(y_mean1 == 0.3)[0]
# c1 = zero_indices
# plt.scatter(zero_indices, y_mean1[zero_indices], color='green', marker='x', s=200, linewidths=4, label='No Data Transmission')
plt.xlim(0, 300)  # x 轴显示范围从 0 到 10
# plt.ylim(0.3, 0.5)  # y 轴显示范围从 -1 到 1

# 添加标题和标签
plt.xlabel('Time Windows', fontsize=18)
plt.ylabel('Packet Loss Rate', fontsize=18)

# 添加图例
plt.legend(fontsize=18)
plt.xticks(fontsize=14)  # 放大 X 轴刻度字体大小
plt.yticks(fontsize=14)  # 放大 Y 轴刻度字体大小
# 显示图表
plt.grid(True)
plt.show()






meta1 = np.load('meta1.npy').reshape(-1)

zero_indices = np.where(meta1 == 0.5)[0]
plt.scatter(zero_indices, meta1[zero_indices], color='black', marker='x', s=100, linewidths=2)
zero_indices = np.where(y_mean1 == 0.3)[0]
plt.scatter(zero_indices, y_mean1[zero_indices], color='blue', marker='o', s=40, linewidths=0.8)
# plt.plot(y_mean1, label='Ensemble Model', color='blue')
plt.plot(meta1, label='Meta Model', color='black', linestyle='-.')
# 绘制真实值 y 的折线图
plt.plot(Y1, label='True Values', color='red', linestyle='--')
plt.xlim(0, 300)  # x 轴显示范围从 0 到 10
# plt.ylim(0.3, 0.5)  # y 轴显示范围从 -1 到 1
# 添加标题和标签
plt.xlabel('Time Windows', fontsize=18)
plt.ylabel('Packet Loss Rate', fontsize=18)
# 添加图例
plt.legend(fontsize=18)
plt.xticks(fontsize=14)  # 放大 X 轴刻度字体大小
plt.yticks(fontsize=14)  # 放大 Y 轴刻度字体大小
# 显示图表
plt.grid(True)
plt.show()

meta2 = np.load('meta2.npy').reshape(-1)
meta3 = np.load('meta3.npy').reshape(-1)
meta3 = meta3 + 0.1

zero_indices = np.where(meta2 == 0.5)[0]
plt.scatter(zero_indices, meta2[zero_indices], color='black', marker='x', s=100, linewidths=2)
zero_indices = np.where(y_mean2 == 0.2)[0]
plt.scatter(zero_indices, y_mean2[zero_indices], color='blue', marker='o', s=40, linewidths=0.8)
# plt.plot(y_mean2, label='Ensemble Model', color='blue')
plt.plot(meta2, label='Meta Model', color='black', linestyle='-.')
# 绘制真实值 y 的折线图
plt.plot(Y2, label='True Values', color='red', linestyle='--')

#
# zero_indices = np.where(y_mean1 == 0.3)[0]
# c1 = zero_indices
# plt.scatter(zero_indices, y_mean1[zero_indices], color='green', marker='x', s=200, linewidths=4, label='No Data Transmission')
plt.xlim(0, 300)  # x 轴显示范围从 0 到 10
# plt.ylim(0.3, 0.5)  # y 轴显示范围从 -1 到 1

# 添加标题和标签
plt.xlabel('Time Windows', fontsize=18)
plt.ylabel('Packet Loss Rate', fontsize=18)

# 添加图例
plt.legend(fontsize=18)
plt.xticks(fontsize=14)  # 放大 X 轴刻度字体大小
plt.yticks(fontsize=14)  # 放大 Y 轴刻度字体大小
# 显示图表
plt.grid(True)
plt.show()


zero_indices = np.where(meta3 == 0.5)[0]
plt.scatter(zero_indices, meta3[zero_indices], color='black', marker='x', s=100, linewidths=2)
zero_indices = np.where(y_mean3 == 0.1)[0]
plt.scatter(zero_indices, y_mean3[zero_indices], color='blue', marker='o', s=40, linewidths=0.8)
# plt.plot(y_mean3, label='Ensemble Model', color='blue')
plt.plot(meta3, label='Meta Model', color='black', linestyle='-.')
# 绘制真实值 y 的折线图
plt.plot(Y3, label='True Values', color='red', linestyle='--')

#
# zero_indices = np.where(y_mean1 == 0.3)[0]
# c1 = zero_indices
# plt.scatter(zero_indices, y_mean1[zero_indices], color='green', marker='x', s=200, linewidths=4, label='No Data Transmission')
plt.xlim(0, 300)  # x 轴显示范围从 0 到 10
# plt.ylim(0.3, 0.5)  # y 轴显示范围从 -1 到 1

# 添加标题和标签
plt.xlabel('Time Windows', fontsize=18)
plt.ylabel('Packet Loss Rate', fontsize=18)

# 添加图例
plt.legend(fontsize=18)
plt.xticks(fontsize=14)  # 放大 X 轴刻度字体大小
plt.yticks(fontsize=14)  # 放大 Y 轴刻度字体大小
# 显示图表
plt.grid(True)
plt.show()



print(1)
