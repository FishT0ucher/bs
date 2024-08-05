import math
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
import statsmodels.api as sm
import torch
import torch.optim.lr_scheduler as lr_scheduler
import xgboost as xgb
from scipy.signal import welch
from scipy.stats import kurtosis, skew
from sklearn.metrics import pairwise_distances
from sklearn.model_selection import train_test_split
from sklearn.neighbors import NearestNeighbors
from statsmodels.tsa.stattools import acf
from statsmodels.tsa.stattools import adfuller, kpss
from torch import nn
from torch.autograd import Variable


def nonlinearfeatures(signal):
    """计算时间序列的特征指标"""
    # 峰度
    kurt = kurtosis(signal)
    # 偏度
    skewness = skew(signal)
    # 非线性灰度关联度
    d = pairwise_distances(signal.reshape(-1, 1))
    max_d = np.max(d)
    norm_d = d / max_d
    epsilon = 0.15 * max_d
    k = 5
    nbrs = NearestNeighbors(n_neighbors=k + 1, algorithm='ball_tree').fit(signal.reshape(-1, 1))
    indices = nbrs.kneighbors(return_distance=False)[:, 1:]
    L = np.zeros_like(norm_d)
    for i in range(signal.shape[0]):
        for j in range(k):
            idx = indices[i, j]
            if idx > i:
                if norm_d[i, idx] < epsilon:
                    L[i, idx] = 1 - (norm_d[i, idx] / epsilon)
                    L[idx, i] = L[i, idx]
    lg = np.sum(L) / (signal.shape[0] * (signal.shape[0] - 1))
    return kurt, skewness, lg,



def spectral_entropy(signal, fs, nperseg, normalize=True):
    f, Pxx = welch(signal, fs=fs, nperseg=nperseg)
    # Normalize the power spectrum
    if normalize:
        Pxx /= np.sum(Pxx)
    # Compute spectral entropy
    spectral_entropy = -np.sum(Pxx * np.log2(Pxx))
    return spectral_entropy


def get_feature(series):
    length = len(series)
    average = np.mean(series)  # 均值
    std = np.std(series)  # 方差
    Max = np.max(series)
    Min = np.min(series)
    # decomposition = seasonal_decompose(series, period=int(length/5))
    # # 得到分解后的趋势、季节性和残差成分
    # trend = decomposition.trend
    # seasonal = decomposition.seasonal
    # residual = decomposition.resid
    # # 计算y的一阶导数
    # y1 = np.gradient(series)
    # # 计算y1的一阶导数
    # # y2 = np.gradient(y1)
    # # # 计算曲率
    # # curvature = np.abs(y2) / np.power(1 + np.square(y1), 1.5)
    # entropy = spectral_entropy(series, 1, length/2, normalize=True)
    # # ADF检验
    # result_adf = adfuller(series)
    # ADF = result_adf[1]
    # # KPSS检验
    # result_kpss = kpss(series)
    # KPSS = result_kpss[1]
    # acf_values = sm.tsa.acf(series)
    # acf_first_value = acf_values[1]
    # # 计算ACF
    # # acf_vals = acf(series, fft=False)
    # # 计算ACF序列前n个值的平方和
    # n = length // 2
    # acf_square_sum = sum(acf_values[1:n + 1] ** 2)
    #
    # # 假设原始序列为y
    # diff_y = np.diff(series)  # 计算一阶差分
    # acf_diff_y = acf(diff_y)  # 计算差分序列的自相关函数
    # first_acf_diff_y = acf_diff_y[1]
    #
    # kurt, skewness, lg = nonlinearfeatures(series)


    return length,average,std,Max,Min


def split_ndarray(ndarray, n):
    segments = len(ndarray) // n
    return [ndarray[i*n:(i+1)*n] for i in range(segments)]



def batch_ave(array):
    batchsize, timestep = array.shape[0], array.shape[1]
    ave = np.full([batchsize, batchsize+timestep-1], fill_value=-100, dtype=np.float32)
    space = 0
    for i in range(batchsize):
        for j in range(timestep):
            ave[i][j+space] = array[i][j]
        space = space + 1
    ave_tensor = np.full([batchsize+timestep-1], fill_value=0, dtype=np.float32)

    for i in range(ave.shape[1]):
        sum = 0
        num = 0
        for j in range(ave.shape[0]):
            if ave[j][i] != -100:
                num = num + 1
                sum = sum +ave[j][i]
        ave_tensor[i] = sum/num
    return ave_tensor





def create_input_list(x_test, n):
    num_samples = len(x_test) // n
    x_test_trimmed = x_test[:num_samples*n]
    x_test_reshaped = np.reshape(x_test_trimmed, (num_samples, n, 1))
    input_list = [torch.from_numpy(x).float() for x in x_test_reshaped]
    return input_list


def expand_and_copy(data_list, n):
    """
    将列表里的每个张量拓展为1，5，1再在第一个维度上复制n个得到n，5，1。
    Args:
    - data_list(list): 包含若干个1，5的张量的列表
    - n(int): 复制的次数
    Return:
    - expanded_list(list): 处理后的列表，包含若干个n，5，1的张量
    """
    expanded_list = []
    for data in data_list:
        # 将每个张量拓展为1，5，1
        expanded_data = data.unsqueeze(-1)
        # 在第一个维度上复制n个得到n，5，1
        expanded_data = torch.cat([expanded_data] * n, dim=0)
        expanded_data = expanded_data.reshape(n, -1, 1)
        expanded_list.append(expanded_data)
    return expanded_list


def split_tensor_by_even_odd(tensor):
    even_tensor = tensor[:, ::2, :]
    odd_tensor = tensor[:, 1::2, :]
    return even_tensor, odd_tensor

def merge_tensors(t1, t2):
    # 切片将t1和t2在第二维度分别拆成奇数位置和偶数位置的张量
    odd_t1 = t1[:, ::2, :]
    even_t2 = t2[:, 1::2, :]
    # 在第二维度上合并
    merged = torch.cat((odd_t1, even_t2), dim=1)
    # 在奇数位置上使用t1，偶数位置上使用t2
    for i in range(merged.size(0)):
        for j in range(merged.size(1)):
            if j % 2 == 0:
                merged[i, j] = t1[i, j//2]
            else:
                merged[i, j] = t2[i, j//2]
    return merged


def reshape_input(batchlist, timestep, dim):
    list2 = []
    for i in range(len(batchlist)):
        batch = create_dataset(batchlist[i], timestep)
        batch = batch.reshape(-1, time_step, dim)
        list2.append(batch)
    return list2

def split_to_batches(data, batch_length):
    num_batches = len(data) // batch_length
    data = data[:num_batches * batch_length] # 舍弃多余部分
    return data.reshape(num_batches, batch_length)


def generate_loss_sequence_Markov(states, transition_matrix, loss_rates, num_packets):
    loss_sequence = []
    current_state = np.random.choice(states)
    for i in range(num_packets):
        loss_prob = loss_rates[current_state]
        loss = np.random.binomial(1, loss_prob)
        loss_sequence.append(loss)
        transition_probs = transition_matrix[current_state]
        current_state = np.random.choice(states, p=transition_probs)
    return loss_sequence

def calculate_loss_rate(loss_sequence, n):
    loss_rates = []
    for i in range(0, len(loss_sequence), n):
        num_packets = min(n, len(loss_sequence) - i)
        num_losses = sum(loss_sequence[i:i+num_packets])
        loss_rate = num_losses / num_packets
        loss_rates.append(loss_rate)
    return loss_rates


def nonsmooth(y, attenuation):
    l = [0]
    l[0] = y[0]
    for i in range(1, len(y)-1, 1):
        l.append(attenuation*y[i]+(1-attenuation)*l[i-1])
    return l

# 以batch形式产生数据集
def create_dataset(dataset, timestep):
    datax = []
    for i in range(0, len(dataset) - timestep + 1, 1):
        batch = dataset[i:(i + timestep)]
        datax.append(batch)
    return np.array(datax)

def separate_dataset(dataset, batch_size):
    datax = []
    for i in range(0, len(dataset)):
        if i % pow(batch_size, 2) == 0:
            for j in range(0, batch_size):
                datax.append(dataset[i+j])
    return np.array(datax)

def ave_normalize(arr):
    normalize_element = []
    arr_new = []
    element = np.average(arr, axis=1)
    element = np.array(element)
    for i in range(len(element)):
        normalize_element.append((element[i]))
    for i in range(len(arr)):
        arr_new.append(list(map(lambda x: math.log(x/element[i]), arr[i])))
    return arr_new, normalize_element

# 归一化
def normalize(arr):
    normalize_element = []
    arr_new = []
    for i in range(0, len(arr), 1):
        element = arr[i][-1]
        normalize_element.append(element)
        arr_new.append(list(map(lambda x: math.log(x/element), arr[i])))
    return arr_new, normalize_element

def normalizey(arr ,normalize_element):
    arr_new = []
    for i in range(0, len(arr), 1):
        arr_new.append(list(map(lambda x: math.log(x/normalize_element[i]), arr[i])))
    return arr_new



# 还原归一化 non seasonal model
def non_normalize(arr, normalize_element):
    arr_new = []
    for i in range(0,len(arr),1):
        arr_new.append(list(map(lambda x: normalize_element[i] * math.exp(x), arr[i])))
    return arr_new

# batch_size是一个batch的长度的意思
def dilation_LSTM(x_train, x_test, y_train, y_test, batch_length, input_length, time_step,hidden_size, num_layer, attenuation, alpha, epoch, lr):
    class d12lstm(nn.Module):
        def __init__(self, input_size, output_size, hidden_size, num_layer, dim):  # 如果你使用8个自变量来预测3个因变量，那么input_size=8，output_size=3
            super(d12lstm, self).__init__()

            self.linear_in = nn.Linear(input_size, input_length)
            self.lstm1 = nn.LSTM(dim, hidden_size, num_layer, batch_first=True)
            # input_size, hidden_size, num_layers  输入数据的维度，输出数据的维度，层数不影响输出。 batchfirst 输入数据是 batchsize，timesstep，dim 前两个由输入决定，dim由hiddensize决定
            self.lstm2_1 = nn.LSTM(hidden_size, hidden_size, num_layer, batch_first=True)
            self.lstm2_2 = nn.LSTM(hidden_size, hidden_size, num_layer, batch_first=True)
            self.linear_out = nn.Linear(input_length*hidden_size, time_step)
            self.tanh = nn.ELU(alpha)

        def forward(self, x):
            bz, tp, di = x.size()
            x = x.view(bz, -1)
            x = self.linear_in(x)
            x = self.tanh(x)
            x = x.view(bz, -1, di)
            x, _ = self.lstm1(x)
            bzx, tpx, dix = x.size()
            lstm2input1, lstm2input2 = split_tensor_by_even_odd(x)
            x21, _ = self.lstm2_1(lstm2input1)
            x22, _ = self.lstm2_2(lstm2input2)
            bz21, tp21, di21 = x21.size()
            bz22, tp22, di22 = x22.size()
            lstm2output = []
            for i in range(bzx):
                if tp21 == tp22:
                    for j in range(tp21):
                        lstm2output.append(x21[i][j])
                        lstm2output.append(x22[i][j])
                else:
                    for j in range(tp22):
                        lstm2output.append(x21[i][j])
                        lstm2output.append(x22[i][j])
                    lstm2output.append(x21[i][tp - 1][-1])
            stacked = torch.stack(lstm2output, dim=0)
            # 沿着新维度进行拼接，得到形状为(6, 8, 4)的张量
            lstm2output = stacked.view(bzx, tpx, -1)
            lstm2output = lstm2output.view(bz, -1)
            y = self.linear_out(lstm2output)
            y = self.tanh(y)
            y = y.view(bz, tp, di)
            return y

    class CustomLoss(nn.Module):
        def __init__(self):
            super(CustomLoss, self).__init__()

        def forward(self, x, y):
            # meanx = torch.mean(x)
            # meany = torch.mean(y)
            # if meanx
            #
            #
            #
            # mse = torch.pow((x - y), 2)
            # mse = torch.mean(mse)
            mse_loss = torch.mean(torch.pow((x - y), 2))  # x与y相减后平方，求均值即为MSE
            return mse_loss

    class CustomLoss2(nn.Module):
        def __init__(self):
            super(CustomLoss2, self).__init__()

        def forward(self, x, y):
            LVP = 0.1
            e = []
            a, b, c = y.size()
            for i in range(a):
                for j in range(1, b-1, 1):
                    y2 = x[i, j+1, c-1].item()
                    y1 = x[i, j-1, c-1].item()
                    e.append(math.log(math.exp(y2)/math.exp(y1)))
            mean_square = sum([x ** 2 for x in e]) / len(e)
            mse_loss = torch.mean(torch.pow((x - y), 2)) + LVP * mean_square # x与y相减后平方，求均值即为MSE
            return mse_loss


    dim = x_train.shape[1]
    x_train = x_train.astype('float32')
    y_train = y_train.astype('float32')
    x_test = x_test.astype('float32')

    x_train = split_to_batches(x_train, batch_length)
    y_train = split_to_batches(y_train, batch_length)
    x_test = split_to_batches(x_test, batch_length)

    y_train = nonsmooth(y_train, attenuation)
    x_train = nonsmooth(x_train, attenuation)

    x_train, normalize_elementx = normalize(x_train)
    y_train = normalizey(y_train, normalize_elementx)
    x_test, normalize_elementt = normalize(x_test)
    # x14, normalize_elementx = ave_normalize(x14)
    # y, normalize_elementy = ave_normalize(y)

    x_train = np.array(x_train)
    y_train = np.array(y_train)
    x_test = np.array(x_test)
    batch_listx = [[row[i] for i in range(len(row))] for row in x_train]
    batch_listy = [[row[i] for i in range(len(row))] for row in y_train]
    batch_listt = [[row[i] for i in range(len(row))] for row in x_test]
    batch_listx = reshape_input(batch_listx, time_step, dim)
    batch_listy = reshape_input(batch_listy, time_step, dim)
    batch_listt = reshape_input(batch_listt, time_step, dim)
    batch_list_tensorx = [torch.Tensor(x) for x in batch_listx]
    batch_list_tensory = [torch.Tensor(x) for x in batch_listy]
    batch_list_tensort = [torch.Tensor(x) for x in batch_listt]

    model = d12lstm(input_size=time_step, output_size=time_step, hidden_size=hidden_size, num_layer=num_layer, dim=dim)
    # criterion = nn.MSELoss()
    criterion = CustomLoss2()
    optimizer = torch.optim.Adam(model.parameters(), lr)
    scheduler = lr_scheduler.StepLR(optimizer, step_size=20, gamma=0.8)
    # 开始训练
    for e in range(epoch):
        # start_time = time.time()
        for batch_num in range(len(batch_list_tensorx)):
            var_x = Variable(batch_list_tensorx[batch_num])
            var_y = Variable(batch_list_tensory[batch_num])
            var_x = var_x.float()
            var_y = var_y.float()
            # 前向传播
            out = model(var_x)
            loss = criterion(out, var_y)
            # 反向传播
            optimizer.zero_grad()
            loss.backward()
            optimizer.step()
        scheduler.step()
        # end_time = time.time()
        # print(end_time - start_time)
        if (e + 1) % 1 == 0:  # 每 100 次输出结果
            print('Epoch: {}, Loss: {:.5f}'.format(e + 1, loss.item()))

    pred_list = []
    model = model.eval()  # 转换成测试模式
    # x_test_list = create_input_list(x_test, time_step)
    # x_test_tensor = expand_and_copy(x_test_list, batch_length-time_step+1)
    for num in range(len(batch_list_tensort)):
        pred_test = model(batch_list_tensort[num])
        pred_test = pred_test.view(-1, time_step).data.numpy()
        pred_test1 = batch_ave(pred_test)
        for i in range(len(pred_test1)):
            pred_test1[i] = math.exp(pred_test1[i]) * normalize_elementt[num]
        pred_list.append(pred_test1)


    pred = np.array(pred_list)
    pred = pred.reshape(-1)

    plt.plot(pred, 'r', label='prediction')
    plt.plot(y_test, 'b', label='real')
    plt.legend(loc='best')
    plt.show()
    return pred_list



file_path1 = 'data1.xlsx'  # 请替换为您的文件路径
workbook1 = openpyxl.load_workbook(file_path1)
sheet1 = workbook1.active

file_path2 = 'data2.xlsx'  # 请替换为您的文件路径
workbook2 = openpyxl.load_workbook(file_path2)
sheet2 = workbook2.active

file_path3 = 'data3.xlsx'  # 请替换为您的文件路径
workbook3 = openpyxl.load_workbook(file_path3)
sheet3 = workbook3.active

delay1 = [cell.value for cell in sheet1['A']]
snr1 = [cell.value for cell in sheet1['B']]
plr1 = [cell.value for cell in sheet1['C']]
delay1 = np.array(delay1[1:]).reshape(-1, 1)
snr1 = np.array(snr1[1:]).reshape(-1, 1)
plr1 = np.array(plr1[1:]).reshape(-1, 1)

delay2 = [cell.value for cell in sheet2['A']]
snr2 = [cell.value for cell in sheet2['B']]
plr2 = [cell.value for cell in sheet2['C']]
delay2 = np.array(delay2[1:]).reshape(-1, 1)
snr2 = np.array(snr2[1:]).reshape(-1, 1)
plr2 = np.array(plr2[1:]).reshape(-1, 1)

delay3 = [cell.value for cell in sheet3['A']]
snr3 = [cell.value for cell in sheet3['B']]
plr3 = [cell.value for cell in sheet3['C']]
plr0 = [cell.value for cell in sheet3['D']]
delay3 = np.array(delay3[1:]).reshape(-1, 1)
snr3 = np.array(snr3[1:]).reshape(-1, 1)
plr3 = np.array(plr3[1:]).reshape(-1, 1)
plr0 = np.array(plr0[1:]).reshape(-1, 1)






states = [0, 1, 2]
transition_matrix = np.array([[0.99, 0.01, 0.00],
                              [0.01, 0.98, 0.01],
                              [0.01, 0.01, 0.98]])
loss_rates = [0.2, 0.3, 0.4]
num_packets = 192000

loss_sequence = generate_loss_sequence_Markov(states, transition_matrix, loss_rates, num_packets)


loss_rate = calculate_loss_rate(loss_sequence, n=100)
loss_rate = np.array(loss_rate)
loss_rate = loss_rate.reshape(-1, 1)
batch_length = 4
time_step = 4
input_length = 10
# x_train, x_test, y_train, y_test = train_test_split(loss_rate[:len(loss_rate)-batch_length], loss_rate[batch_length:], test_size=0.3, shuffle=False)
# y_list = split_ndarray(y_test, batch_length)


snr1[snr1 < 0] = 16
plr0[plr0 < 0] = 0.5
snr1 =snr1/32
x_train = snr1[700:]
y_train = plr0[700:]
x_test = snr1[700:]
y_test = plr0[700:]
y_list = split_ndarray(y_test, batch_length)


LSTM_prelist = dilation_LSTM(x_train, x_test, y_train, y_test, batch_length, input_length, time_step, hidden_size=3, num_layer=1, attenuation=0.75, alpha=0.5, epoch=10, lr=0.008)

LSTM_prelist2 = dilation_LSTM(x_train, x_test, y_train, y_test, batch_length, input_length, time_step, hidden_size=2, num_layer=1, attenuation=1, alpha=1, epoch=10, lr=0.01)

prelist3 = split_ndarray(x_test, batch_length)

for i in range(len(y_list)):
    y_list[i] = y_list[i].reshape(-1)


l2error = []
weight1 = []
weight2 = []

lower_bound = -1
uper_bound = 1.5
step = 0.01
for i in range(len(y_list)):
    weight1tem = []
    weight2tem = []
    l2errortem = []
    errortem = 100
    for w1 in np.arange(lower_bound, uper_bound, step):
        for w2 in np.arange(lower_bound, uper_bound, step):
            diff = y_list[i] - w1 * LSTM_prelist[i] - w2 * LSTM_prelist2[i]
            l2_norm = np.linalg.norm(diff, ord=2)
            weight1tem.append(w1)
            weight2tem.append(w2)
            l2errortem.append(l2_norm)
    for j in range(len(weight1tem)):
        if errortem > l2errortem[j]:
            errortem = l2errortem[j]
    for j in range(len(weight1tem)):
        if l2errortem[j] == errortem:
            l2error.append(errortem)
            weight1.append(weight1tem[j])
            weight2.append(weight2tem[j])
            break

feature = pd.DataFrame()
for i, ts in enumerate(y_list):
    length, average, std, Max, Min = get_feature(y_list[i])
    feature.loc[i, 'length'] = length
    feature.loc[i, 'mean'] = average
    feature.loc[i, 'std'] = std
    feature.loc[i, 'max'] = Max
    feature.loc[i, 'min'] = Min
    # feature.loc[i, 'trend'] = trend
    # feature.loc[i, 'seasonal'] = seasonal
    # feature.loc[i, 'residual'] = residual
    # feature.loc[i, 'curvature'] = curvature

    # feature.loc[i, 'entropy'] = entropy
    # feature.loc[i, 'ADF'] = ADF
    # feature.loc[i, 'KPSS'] = KPSS
    # feature.loc[i, 'acf_first_value'] = acf_first_value
    # feature.loc[i, 'acf_square_sum'] = acf_square_sum
    # feature.loc[i, 'first_acf_diff_y'] = first_acf_diff_y
    # feature.loc[i, 'kurtosis'] = kurt
    # feature.loc[i, 'skewness'] = skewness
    # feature.loc[i, 'gray_scale'] = gray_scale

weight1 = pd.DataFrame(weight1, columns=['weight1'])
weight2 = pd.DataFrame(weight2, columns=['weight2'])


reg = xgb.XGBRegressor(booster='dart',
                       n_estimators=200,
                       subsample=0.75,
                       # gamma=3,
                       # colsample_bytree=0.75,
                       # colsample_bylevel=0.75,
                       # colsample_bynode=0.75,
                       # min_child_weight=1,
                       reg_alpha=1,
                       reg_lambda=1,
                       objective='reg:squarederror',
                       eval_metric='rmse',
                       n_jobs=-1,
                       max_depth=18,
                       learning_rate=0.02)
reg.fit(feature, weight1, eval_set=[(feature, weight1), (feature, weight1)], verbose=1)
weight1_pred = reg.predict(feature)


reg2 = xgb.XGBRegressor(booster='dart',
                       n_estimators=200,
                       subsample=0.75,
                       # gamma=3,
                       # colsample_bytree=0.75,
                       # colsample_bylevel=0.75,
                       # colsample_bynode=0.75,
                       # min_child_weight=1,
                       reg_alpha=1,
                       reg_lambda=1,
                       objective='reg:squarederror',
                       eval_metric='rmse',
                       n_jobs=-1,
                       max_depth=18,
                       learning_rate=0.02)
reg2.fit(feature, weight2, eval_set=[(feature, weight2), (feature, weight2)], verbose=1)
weight2_pred = reg2.predict(feature)





# plt.plot(weight1_pred, 'r', label='prediction')
# plt.plot(weight1, 'b', label='real')
# plt.legend(loc='best')
# plt.show()

metapred = []
for i in range(len(LSTM_prelist)):
    metapred.append(LSTM_prelist[i]*weight1_pred[i]+LSTM_prelist2[i]*weight2_pred[i])


pred2 = np.array(metapred)
pred2 = pred2.reshape(-1)/1.1

pred2[pred2 == 0.5] = -1
y_test[y_test == 0.5] = -1

plt.plot(pred2, 'r', label='prediction')
plt.plot(y_test, 'b', label='real')
plt.legend(loc='best')
plt.show()

np.save('meta0.npy', pred2)



print(w1, '  : ', l2_norm)



